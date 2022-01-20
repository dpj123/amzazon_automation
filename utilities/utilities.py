import softest
from openpyxl import load_workbook
from selenium.webdriver import Keys


class Utilities(softest.TestCase):

    def autocomplete_list_selection(self, item_list, item):
        for x in item_list:
            print(x.text + item)
            if item == x.text:
                x.click()
                break
            else:
                print("No result found")

    def custom_drop_down_selection(self, selection_list, selected_item):
        for x in selection_list:
            if selected_item == x.text:
                return x.text
                break


    def select_item_from_the_list(self, item_list, selected_item):
        for x in item_list:
            if selected_item in x.text:
                return x

    def soft_assert_the_item(self, item_list, selected_item):
        for x in item_list:
            print(x.text)
            self.soft_assert(self.assertEqual, x.text, selected_item)
            if selected_item == x.text:
                x.click()
            else:
                print("test case failed")
        self.assert_all()

    def assert_the_items(self, item_list, selected_item):
        for x in item_list:
            print(x.text)
            assert selected_item in x.text, "Test case failed"
            print("Test case passed")

    def window_handle(self, parent, child):
        for x in child:
            if x != parent:
                return x

    def read_data_from_excelfile(self, file_name, sheet_name):
        wb = load_workbook(file_name)
        sh = wb[sheet_name]

        row_count = sh.max_row
        column_count = sh.max_column
        print(row_count, column_count)
        data_list = []
        for i in range(2, row_count + 1):
            row_data_store = []
            for j in range(1, column_count + 1):
                store = sh.cell(row=i, column=j).value
                row_data_store.append(store)
            data_list.append(row_data_store)
        print(data_list)
        return data_list


